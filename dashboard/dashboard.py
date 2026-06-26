#!/usr/bin/env python3
"""Conversion dashboard for the Copado / BAU regression suite.

Live local web app that computes conversion status from GROUND TRUTH on every
request — nothing is stored or hand-maintained:

  * Generated     — which workbook test cases now exist in tests/*.robot
  * Review burden — count of unresolved "# CONFIRM IN CRT:" flags per test
  * CRT pass/fail — parsed from the newest Robot output.xml (if present)
  * Coverage gaps — what's still Not started, broken out by tab (+ priority)

Run from this repo's root (it auto-discovers the workbook and tests/):

    python3 dashboard/dashboard.py
    # then open http://localhost:8765  — every reload re-scans the repo

Options: --repo PATH  --workbook PATH  --port N  --host H
The workbook enumeration is cached by file mtime; the .robot files and result
files are re-scanned on every page load.

This dashboard is specific to this project. It reuses one generic component — the
spreadsheet parser `extract_testcase.py` from the installed `copado-regression-skills`
plugin (or a sibling checkout of that repo) — and discovers it at startup.
"""
from __future__ import annotations

import argparse
import glob
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from html import escape
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Dict, List, Optional


def _load_parser():
    """Locate extract_testcase.py from the copado-regression-skills plugin.

    Search order: already-importable -> $NEWTEST_PARSER_DIR -> $CLAUDE_PLUGIN_ROOT ->
    a sibling `copado-regression-skills` checkout near this repo. Exits with a clear
    message if the plugin can't be found.
    """
    try:
        import extract_testcase as _ex  # already on sys.path
        return _ex
    except ImportError:
        pass
    rel = Path("plugins") / "copado-regression-skills" / "skills" / "newtest"
    candidates = []
    if os.environ.get("NEWTEST_PARSER_DIR"):
        candidates.append(Path(os.environ["NEWTEST_PARSER_DIR"]))
    if os.environ.get("CLAUDE_PLUGIN_ROOT"):
        candidates.append(Path(os.environ["CLAUDE_PLUGIN_ROOT"]) / "skills" / "newtest")
    here = Path(__file__).resolve()
    for up in here.parents:
        candidates.append(up / "copado-regression-skills" / rel)
        candidates.append(up / rel)
    for cand in candidates:
        if (cand / "extract_testcase.py").is_file():
            sys.path.insert(0, str(cand))
            import extract_testcase as _ex
            return _ex
    sys.exit(
        "Could not find extract_testcase.py. Install the copado-regression-skills plugin, "
        "or set NEWTEST_PARSER_DIR to the folder containing extract_testcase.py."
    )


ex = _load_parser()

# Functional tabs we convert, mapped to their target .robot file. Order = conversion order.
TABS = {
    "1. Sales": "tests/Sales.robot",
    "2. Case": "tests/Case.robot",
    "3. Genesys Sales": "tests/GenesysSales.robot",
    "4. Genesys Service": "tests/GenesysService.robot",
    "5.LCS": "tests/LCS.robot",
}
# Known case counts — used only as a sanity check printed at startup.
EXPECTED_COUNTS = {"1. Sales": 108, "2. Case": 91, "3. Genesys Sales": 14,
                   "4. Genesys Service": 15, "5.LCS": 6}

TC_RE = re.compile(r"TC[_]?\d+", re.IGNORECASE)
CONFIRM_RE = re.compile(r"#\s*CONFIRM\s+IN\s+CRT", re.IGNORECASE)


def _norm_id(value: str) -> str:
    """Normalise a TC token: 'tc102' / 'TC_102' -> 'TC_102'."""
    m = TC_RE.search(value or "")
    if not m:
        return ""
    num = re.search(r"\d+", m.group(0)).group(0)
    return f"TC_{num}"


# --------------------------------------------------------------------------- #
# 1. Enumerate every case from the workbook (single streaming pass per tab).   #
# --------------------------------------------------------------------------- #
def enumerate_workbook(workbook: Path) -> List[Dict]:
    import openpyxl
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    cases: List[Dict] = []
    try:
        for tab, target in TABS.items():
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]
            header_row = ex.find_header_row(ws)
            cols = ex.map_columns(ws, header_row)
            id_col = cols.get("identity")
            step_col = cols.get("step")
            pr_col = cols.get("priority")
            nm_col = cols.get("name")
            comp_col = cols.get("component")
            if not id_col:
                continue

            def cell(row, idx):
                if not idx or idx - 1 >= len(row):
                    return None
                return row[idx - 1].value

            seen: Dict[str, int] = {}
            cur: Optional[Dict] = None
            for row in ws.iter_rows(min_row=header_row + 1):
                ident = ex._norm(cell(row, id_col))
                if ident:
                    tcid = _norm_id(ident)
                    # disambiguate duplicate ids within a tab with a row suffix
                    key = tcid or ident
                    seen[key] = seen.get(key, 0) + 1
                    start = row[0].row if row else header_row + 1
                    cur = {
                        "tab": tab,
                        "target_file": target,
                        "identity": ident,
                        "tcid": tcid,            # "" for name-based tabs (Genesys/LCS)
                        "name": ex._norm(cell(row, nm_col)),
                        "priority": ex._norm(cell(row, pr_col)),
                        "component": ex._norm(cell(row, comp_col)),
                        "start_row": start,
                        "steps": 0,
                        "occurrence": seen[key],
                    }
                    cases.append(cur)
                if cur is not None and ex._norm(cell(row, step_col)):
                    cur["steps"] += 1
    finally:
        wb.close()
    return cases


# --------------------------------------------------------------------------- #
# 2. Scan tests/*.robot for converted ids + CONFIRM-flag burden.               #
# --------------------------------------------------------------------------- #
def scan_robot_files(repo: Path) -> Dict:
    """Scan tests/*.robot. Conversion is scoped PER FILE (TC ids are reused across
    tabs, so a global match would cross-contaminate).

    Returns:
      converted      {(filename, tcid): {confirm:int}}   — tcid is an actual *test case*
      confirm_in     {(filename, tcid): int}             — flags inside that test's keywords
      names_by_file  {filename: [test-case names]}        — for name-based (Genesys/LCS) tabs
      shared_confirm {filename: int}                      — flags not attributable to a test
    """
    testcases = set()          # {(filename, tcid)} — actual *** Test Cases *** entries
    confirm: Dict[tuple, int] = {}   # {(filename, tcid): flag count}
    names_by_file: Dict[str, List[str]] = {}
    shared_confirm: Dict[str, int] = {}
    for path in sorted((repo / "tests").glob("*.robot")):
        fname = path.name
        names_by_file.setdefault(fname, [])
        text = path.read_text(encoding="utf-8", errors="replace")
        section = None
        block_tc = ""          # TC id of the current test-case / keyword block
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("***"):
                low = stripped.lower()
                if "test case" in low:
                    section = "tests"
                elif "keyword" in low:
                    section = "keywords"
                else:
                    section = None
                block_tc = ""
                continue
            # A block header is a non-indented, non-comment, non-empty line.
            is_header = bool(line) and not line[0].isspace() and not stripped.startswith("#")
            if is_header and section in ("tests", "keywords"):
                block_tc = _norm_id(stripped)
                if section == "tests":
                    names_by_file[fname].append(stripped)
                    if block_tc:
                        testcases.add((fname, block_tc))
            if CONFIRM_RE.search(line):
                if block_tc:
                    confirm[(fname, block_tc)] = confirm.get((fname, block_tc), 0) + 1
                else:
                    shared_confirm[fname] = shared_confirm.get(fname, 0) + 1
    return {"testcases": testcases, "confirm": confirm,
            "names_by_file": names_by_file, "shared_confirm": shared_confirm}


# --------------------------------------------------------------------------- #
# 3. Parse the newest Robot output.xml for pass/fail (optional).               #
# --------------------------------------------------------------------------- #
def parse_results(repo: Path) -> Dict[str, str]:
    candidates = []
    for pat in ("output.xml", "results/**/output.xml", "**/output.xml"):
        candidates += glob.glob(str(repo / pat), recursive=True)
    if not candidates:
        return {}
    newest = max(candidates, key=os.path.getmtime)
    results: Dict[str, str] = {}
    try:
        tree = ET.parse(newest)
    except Exception:
        return {}
    for test in tree.iter("test"):
        name = test.get("name", "")
        tcid = _norm_id(name)
        if not tcid:
            continue
        status_el = test.find("status")
        status = (status_el.get("status") if status_el is not None else "") or ""
        # worst-case wins if an id appears more than once
        if results.get(tcid) == "FAIL":
            continue
        results[tcid] = status.upper()
    return {"_source": os.path.relpath(newest, repo), **results}


# --------------------------------------------------------------------------- #
# Combine everything into the status model.                                    #
# --------------------------------------------------------------------------- #
_WB_CACHE: Dict = {}


def get_cases(workbook: Path) -> List[Dict]:
    mtime = workbook.stat().st_mtime
    key = (str(workbook), mtime)
    if _WB_CACHE.get("key") != key:
        _WB_CACHE["key"] = key
        _WB_CACHE["cases"] = enumerate_workbook(workbook)
    return _WB_CACHE["cases"]


def build_model(repo: Path, workbook: Path) -> Dict:
    cases = get_cases(workbook)
    scan = scan_robot_files(repo)
    testcases = scan["testcases"]
    confirm_map = scan["confirm"]
    names_by_file = scan["names_by_file"]
    results = parse_results(repo)
    result_source = results.pop("_source", None)

    rows = []
    for c in cases:
        tcid = c["tcid"]
        fname = os.path.basename(c["target_file"])
        is_conv = False
        confirm = 0
        # Credit conversion only to the FIRST occurrence of a duplicated id within a tab
        # (the workbook reuses some ids — e.g. Sales TC_001 x6 — for distinct scenarios;
        # the later occurrences are separate cases that need their own row-selector test).
        if tcid and (fname, tcid) in testcases and c["occurrence"] == 1:
            is_conv = True
            confirm = confirm_map.get((fname, tcid), 0)
        elif not tcid and c["name"]:
            # name-based tab (Genesys/LCS): best-effort substring match within the target file
            nm = c["name"].lower()[:25]
            if nm and any(nm in rn.lower() for rn in names_by_file.get(fname, [])):
                is_conv = True
        crt = results.get(tcid, "") if tcid else ""
        if crt == "PASS":
            status = "Passing"
        elif crt == "FAIL":
            status = "Failing"
        elif is_conv and confirm == 0:
            status = "Ready"
        elif is_conv:
            status = "Generated"
        else:
            status = "Not started"
        rows.append({
            "tab": c["tab"], "id": c["identity"], "tcid": tcid,
            "name": c["name"], "priority": c["priority"], "steps": c["steps"],
            "target": c["target_file"], "converted": is_conv,
            "confirm": confirm, "crt": crt, "status": status,
        })

    by_tab: Dict[str, Dict] = {}
    for r in rows:
        t = by_tab.setdefault(r["tab"], {"total": 0, "converted": 0, "ready": 0,
                                         "passing": 0, "failing": 0, "confirm": 0,
                                         "target": r["target"]})
        t["total"] += 1
        t["confirm"] += r["confirm"]
        if r["converted"]:
            t["converted"] += 1
        if r["status"] == "Ready":
            t["ready"] += 1
        if r["status"] == "Passing":
            t["passing"] += 1
        if r["status"] == "Failing":
            t["failing"] += 1

    total = len(rows)
    summary = {
        "total": total,
        "converted": sum(1 for r in rows if r["converted"]),
        "passing": sum(1 for r in rows if r["status"] == "Passing"),
        "failing": sum(1 for r in rows if r["status"] == "Failing"),
        "confirm": sum(r["confirm"] for r in rows),
        "not_started": sum(1 for r in rows if r["status"] == "Not started"),
        "result_source": result_source,
        "shared_confirm": scan["shared_confirm"],
    }
    return {"summary": summary, "by_tab": by_tab, "rows": rows}


# --------------------------------------------------------------------------- #
# HTML rendering                                                               #
# --------------------------------------------------------------------------- #
BADGE = {
    "Not started": "#9aa0a6", "Generated": "#e8a33d", "Ready": "#3b82f6",
    "Passing": "#16a34a", "Failing": "#dc2626",
}


def render_html(model: Dict, repo: Path) -> str:
    s = model["summary"]
    pct = lambda n, d: round(100 * n / d) if d else 0
    conv_pct = pct(s["converted"], s["total"])
    pass_pct = pct(s["passing"], s["total"])

    tab_cards = []
    for tab, t in model["by_tab"].items():
        cp = pct(t["converted"], t["total"])
        pp = pct(t["passing"], t["total"])
        tab_cards.append(f"""
        <div class="card">
          <div class="card-h"><b>{escape(tab)}</b><span class="mono">{escape(t['target'])}</span></div>
          <div class="bar"><div class="fill conv" style="width:{cp}%"></div>
                           <div class="fill pass" style="width:{pp}%"></div></div>
          <div class="stats">
            <span>{t['converted']}/{t['total']} converted ({cp}%)</span>
            <span class="g">{t['passing']} passing</span>
            <span class="r">{t['failing']} failing</span>
            <span class="o">{t['confirm']} flags</span>
          </div>
        </div>""")

    rows_html = []
    for r in sorted(model["rows"], key=lambda x: (list(TABS).index(x["tab"]) if x["tab"] in TABS else 9, x["id"])):
        color = BADGE.get(r["status"], "#9aa0a6")
        crt = r["crt"] or "—"
        rows_html.append(f"""
        <tr data-status="{r['status']}" data-tab="{escape(r['tab'])}">
          <td class="mono">{escape(r['tab'])}</td>
          <td class="mono">{escape(r['id'])}</td>
          <td>{escape(r['name'][:70])}</td>
          <td class="num">{r['steps']}</td>
          <td>{escape(r['priority'] or '')}</td>
          <td><span class="badge" style="background:{color}">{r['status']}</span></td>
          <td class="num">{r['confirm'] or ''}</td>
          <td class="mono">{crt}</td>
        </tr>""")

    # Coverage gaps — Not started grouped by tab, then priority.
    gaps = {}
    for r in model["rows"]:
        if r["status"] == "Not started":
            gaps.setdefault(r["tab"], []).append(r)
    gap_html = []
    for tab in TABS:
        items = gaps.get(tab, [])
        if not items:
            continue
        by_pri = {}
        for r in items:
            by_pri.setdefault(r["priority"] or "—", []).append(r["id"])
        pri_bits = " · ".join(f"{escape(p)}: {len(ids)}" for p, ids in sorted(by_pri.items()))
        nxt = ", ".join(escape(r["id"]) for r in items[:6])
        gap_html.append(f"<li><b>{escape(tab)}</b> — {len(items)} left ({pri_bits})<br>"
                        f"<span class='mono next'>next: {nxt}{' …' if len(items) > 6 else ''}</span></li>")

    # Tab bar — one button per functional tab (+ All) to slice the table by tab.
    tab_buttons = ['<button class="tab active" data-tab="" onclick="pickTab(this)">All</button>']
    for tab in TABS:
        n_tab = sum(1 for r in model["rows"] if r["tab"] == tab)
        tab_buttons.append(
            f'<button class="tab" data-tab="{escape(tab)}" onclick="pickTab(this)">'
            f'{escape(tab)} <span class="cnt">{n_tab}</span></button>')

    src = s["result_source"]
    src_line = (f"CRT results: <span class='mono'>{escape(src)}</span>"
                if src else "CRT results: <i>none found (run a suite to populate pass/fail)</i>")

    return f"""<!doctype html><html><head><meta charset="utf-8">
<title>Regression Conversion Dashboard</title>
<style>
  :root {{ color-scheme: light dark; }}
  body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 0; background:#0f1115; color:#e6e6e6; }}
  .wrap {{ max-width: 1080px; margin: 0 auto; padding: 24px; }}
  h1 {{ font-size: 20px; margin: 0 0 4px; }}
  .sub {{ color:#9aa0a6; font-size:13px; margin-bottom:18px; }}
  .hero {{ display:flex; gap:18px; margin-bottom:20px; flex-wrap:wrap; }}
  .kpi {{ background:#1a1d24; border:1px solid #2a2e37; border-radius:12px; padding:14px 18px; min-width:120px; }}
  .kpi .n {{ font-size:26px; font-weight:700; }}
  .kpi .l {{ font-size:12px; color:#9aa0a6; }}
  .bar {{ position:relative; height:10px; background:#2a2e37; border-radius:6px; overflow:hidden; margin:8px 0; }}
  .fill {{ position:absolute; top:0; left:0; height:100%; }}
  .fill.conv {{ background:#3b82f6; opacity:.5; }}
  .fill.pass {{ background:#16a34a; }}
  .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; }}
  .card {{ background:#1a1d24; border:1px solid #2a2e37; border-radius:12px; padding:12px 14px; }}
  .card-h {{ display:flex; justify-content:space-between; font-size:13px; margin-bottom:4px; }}
  .stats {{ font-size:12px; color:#c8ccd2; display:flex; gap:12px; flex-wrap:wrap; }}
  .stats .g {{ color:#4ade80; }} .stats .r {{ color:#f87171; }} .stats .o {{ color:#fbbf24; }}
  .mono {{ font-family: ui-monospace, Menlo, monospace; color:#9aa0a6; font-size:12px; }}
  table {{ width:100%; border-collapse:collapse; margin-top:8px; font-size:13px; }}
  th, td {{ text-align:left; padding:6px 8px; border-bottom:1px solid #23272f; }}
  th {{ color:#9aa0a6; font-weight:600; font-size:11px; text-transform:uppercase; letter-spacing:.04em; }}
  td.num {{ text-align:right; }}
  .badge {{ color:#fff; padding:2px 8px; border-radius:20px; font-size:11px; font-weight:600; }}
  .next {{ color:#7f8896; }}
  .toolbar {{ display:flex; gap:8px; align-items:center; margin:18px 0 4px; flex-wrap:wrap; }}
  button, select {{ background:#1a1d24; color:#e6e6e6; border:1px solid #2a2e37; border-radius:8px; padding:6px 12px; font-size:13px; cursor:pointer; }}
  .tabbar {{ display:flex; gap:6px; margin:22px 0 0; flex-wrap:wrap; border-bottom:1px solid #2a2e37; }}
  .tab {{ background:transparent; border:1px solid transparent; border-bottom:none; border-radius:8px 8px 0 0;
          color:#9aa0a6; padding:8px 14px; font-size:13px; cursor:pointer; margin-bottom:-1px; }}
  .tab:hover {{ color:#e6e6e6; }}
  .tab.active {{ background:#1a1d24; border-color:#2a2e37; border-bottom-color:#1a1d24; color:#e6e6e6; font-weight:600; }}
  .tab .cnt {{ color:#7f8896; font-size:11px; }}
  .section-h {{ font-size:14px; margin:22px 0 6px; }}
  ul {{ line-height:1.6; }}
  a {{ color:#60a5fa; }}
</style></head><body><div class="wrap">
  <h1>Regression Conversion Dashboard</h1>
  <div class="sub">Live from <span class="mono">{escape(str(repo))}</span> — recomputed on every reload. {src_line}</div>

  <div class="hero">
    <div class="kpi"><div class="n">{s['converted']}/{s['total']}</div><div class="l">converted ({conv_pct}%)</div></div>
    <div class="kpi"><div class="n" style="color:#4ade80">{s['passing']}</div><div class="l">passing in CRT ({pass_pct}%)</div></div>
    <div class="kpi"><div class="n" style="color:#f87171">{s['failing']}</div><div class="l">failing</div></div>
    <div class="kpi"><div class="n" style="color:#fbbf24">{s['confirm']}</div><div class="l">open CONFIRM flags</div></div>
    <div class="kpi"><div class="n">{s['not_started']}</div><div class="l">not started</div></div>
  </div>
  <div class="bar" style="height:14px"><div class="fill conv" style="width:{conv_pct}%"></div><div class="fill pass" style="width:{pass_pct}%"></div></div>

  <div class="section-h">By tab</div>
  <div class="grid">{''.join(tab_cards)}</div>

  <div class="section-h">What's next — coverage gaps</div>
  <ul>{''.join(gap_html) or '<li>Nothing left — every case is converted.</li>'}</ul>

  <div class="tabbar" id="tabbar">{''.join(tab_buttons)}</div>
  <div class="toolbar">
    <button onclick="location.reload()">↻ Re-scan</button>
    <label>Filter status
      <select id="f" onchange="filt()">
        <option value="">all</option>
        <option>Not started</option><option>Generated</option><option>Ready</option>
        <option>Passing</option><option>Failing</option>
      </select></label>
    <span class="mono" id="count"></span>
  </div>
  <table id="t"><thead><tr>
    <th>Tab</th><th>ID</th><th>Name</th><th>Steps</th><th>Priority</th><th>Status</th><th>Flags</th><th>CRT</th>
  </tr></thead><tbody>{''.join(rows_html)}</tbody></table>
</div>
<script>
  var curTab = '';
  function pickTab(btn) {{
    curTab = btn.dataset.tab;
    document.querySelectorAll('#tabbar .tab').forEach(function(b) {{ b.classList.toggle('active', b === btn); }});
    filt();
  }}
  function filt() {{
    var v = document.getElementById('f').value, n = 0;
    document.querySelectorAll('#t tbody tr').forEach(function(tr) {{
      var show = (!v || tr.dataset.status === v) && (!curTab || tr.dataset.tab === curTab);
      tr.style.display = show ? '' : 'none'; if (show) n++;
    }});
    document.getElementById('count').textContent = n + ' shown';
  }}
  filt();
</script>
</body></html>"""


# --------------------------------------------------------------------------- #
# Server                                                                       #
# --------------------------------------------------------------------------- #
def make_handler(repo: Path, workbook: Path):
    class Handler(BaseHTTPRequestHandler):
        def log_message(self, *a):  # quiet
            pass

        def do_GET(self):
            try:
                model = build_model(repo, workbook)
            except Exception as e:  # surface errors in the page rather than 500-ing silently
                body = f"<pre>Dashboard error:\n{escape(repr(e))}</pre>".encode()
                self.send_response(500); self.send_header("Content-Type", "text/html"); self.end_headers()
                self.wfile.write(body); return
            if self.path.startswith("/api"):
                payload = json.dumps(model, indent=2).encode()
                self.send_response(200); self.send_header("Content-Type", "application/json")
                self.end_headers(); self.wfile.write(payload); return
            html = render_html(model, repo).encode()
            self.send_response(200); self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers(); self.wfile.write(html)
    return Handler


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Live regression-conversion dashboard.")
    ap.add_argument("--repo", default=os.getcwd(), help="robot-files repo root (default: cwd)")
    ap.add_argument("--workbook", default=None, help="explicit path to the .xlsx")
    ap.add_argument("--port", type=int, default=8765)
    ap.add_argument("--host", default="127.0.0.1")
    ap.add_argument("--once", action="store_true", help="print a one-shot text report and exit")
    args = ap.parse_args(argv)

    repo = Path(args.repo).resolve()
    workbook = Path(args.workbook) if args.workbook else ex.find_workbook(repo)
    if not Path(workbook).is_file():
        print(f"Workbook not found (looked from {repo}). Pass --workbook.", file=sys.stderr)
        return 2

    # startup sanity check
    cases = get_cases(Path(workbook))
    counts = {}
    for c in cases:
        counts[c["tab"]] = counts.get(c["tab"], 0) + 1
    print(f"[dashboard] workbook: {workbook}", file=sys.stderr)
    for tab, exp in EXPECTED_COUNTS.items():
        got = counts.get(tab, 0)
        flag = "" if got == exp else f"  <-- expected {exp}"
        print(f"[dashboard]   {tab}: {got} cases{flag}", file=sys.stderr)

    if args.once:
        model = build_model(repo, Path(workbook))
        s = model["summary"]
        print(f"\nConverted {s['converted']}/{s['total']} | passing {s['passing']} | "
              f"failing {s['failing']} | open flags {s['confirm']} | not started {s['not_started']}")
        for tab, t in model["by_tab"].items():
            print(f"  {tab:18} {t['converted']:>3}/{t['total']:<3} converted, "
                  f"{t['passing']} pass, {t['confirm']} flags")
        return 0

    httpd = ThreadingHTTPServer((args.host, args.port), make_handler(repo, Path(workbook)))
    print(f"[dashboard] serving http://{args.host}:{args.port}  (Ctrl-C to stop)", file=sys.stderr)
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n[dashboard] stopped.", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
